# coding: utf-8
"""
Script: excel_exporter.py
Função: Exportador para formato Excel
Autor: Nicolas Avila
Data: 2025-11-11
Projeto: Avila Ops - Windows Dev Optimizer Framework
Descrição: Exportação de relatórios em formato Excel
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import Dict, Any, List
import json

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    pd = None
    Workbook = None

logger = logging.getLogger(__name__)

class ExcelExporter:
    """Exportador para formato Excel"""
    
    def __init__(self, output_dir: str = "output"):
        if not EXCEL_AVAILABLE:
            logger.warning("Bibliotecas Excel não disponíveis. Instale: pip install pandas openpyxl")
        
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
    
    async def export_analysis_report(self, analysis_data: Dict[str, Any], filename: str = None) -> str:
        """Exporta relatório de análise em Excel"""
        
        if not EXCEL_AVAILABLE:
            raise ImportError("Bibliotecas Excel não disponíveis. Instale: pip install pandas openpyxl")
        
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"windows_analysis_report_{timestamp}.xlsx"
        
        filepath = self.output_dir / filename
        
        try:
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                
                # Aba 1: Resumo
                self._create_summary_sheet(analysis_data, writer)
                
                # Aba 2: Informações do Sistema
                if "system_info" in analysis_data and "error" not in analysis_data.get("system_info", {}):
                    self._create_system_info_sheet(analysis_data["system_info"], writer)
                
                # Aba 3: Análise de Programas
                if "programs_analysis" in analysis_data and "error" not in analysis_data.get("programs_analysis", {}):
                    self._create_programs_sheet(analysis_data["programs_analysis"], writer)
                
                # Aba 4: Edge Analysis
                if "edge_analysis" in analysis_data and "error" not in analysis_data.get("edge_analysis", {}):
                    self._create_edge_sheet(analysis_data["edge_analysis"], writer)
                
                # Aba 5: Bloatware
                if "bloatware_detected" in analysis_data:
                    self._create_bloatware_sheet(analysis_data["bloatware_detected"], writer)
            
            # Aplicar formatação
            self._apply_formatting(filepath)
            
            logger.info(f"Relatório Excel exportado para: {filepath}")
            return str(filepath)
            
        except Exception as e:
            logger.error(f"Erro ao exportar relatório Excel: {e}")
            raise
    
    def _create_summary_sheet(self, data: Dict[str, Any], writer):
        """Cria aba de resumo"""
        
        summary_data = []
        timestamp = data.get("timestamp", datetime.now().isoformat())
        
        summary_data.append(["Windows Dev Optimizer - Relatório de Análise", ""])
        summary_data.append(["Data de Geração", timestamp])
        summary_data.append(["", ""])
        
        # Estatísticas gerais
        summary_data.append(["RESUMO EXECUTIVO", ""])
        summary_data.append(["", ""])
        
        # Sistema
        if "system_info" in data:
            sys_info = data["system_info"]
            if "error" not in sys_info:
                summary_data.append(["Memória Total (GB)", sys_info.get("memory_total_gb", "N/A")])
                summary_data.append(["Memória Disponível (GB)", sys_info.get("memory_available_gb", "N/A")])
        
        # Programas
        if "programs_analysis" in data:
            prog_analysis = data["programs_analysis"]
            if "error" not in prog_analysis:
                summary_data.append(["Total de Programas", prog_analysis.get("total_programs", "N/A")])
                if "unused_programs" in prog_analysis:
                    summary_data.append(["Programas Não Utilizados", len(prog_analysis["unused_programs"])])
        
        # Bloatware
        if "bloatware_detected" in data:
            summary_data.append(["Bloatware Detectado", len(data["bloatware_detected"])])
        
        # Edge
        if "edge_analysis" in data and "history_stats" in data["edge_analysis"]:
            edge_stats = data["edge_analysis"]["history_stats"]
            summary_data.append(["Sites Únicos Visitados", edge_stats.get("unique_sites", "N/A")])
        
        df_summary = pd.DataFrame(summary_data, columns=["Métrica", "Valor"])
        df_summary.to_excel(writer, sheet_name="Resumo", index=False)
    
    def _create_system_info_sheet(self, system_info: Dict[str, Any], writer):
        """Cria aba de informações do sistema"""
        
        # Informações básicas
        basic_info = []
        basic_info.append(["Plataforma", system_info.get("platform", "N/A")])
        basic_info.append(["Processador", system_info.get("processor", "N/A")])
        basic_info.append(["Arquitetura", system_info.get("architecture", ["N/A"])[0]])
        basic_info.append(["Memória Total (GB)", system_info.get("memory_total_gb", "N/A")])
        basic_info.append(["Memória Disponível (GB)", system_info.get("memory_available_gb", "N/A")])
        
        df_basic = pd.DataFrame(basic_info, columns=["Especificação", "Valor"])
        
        # Uso de disco
        disk_info = []
        if "disk_usage" in system_info:
            for drive, info in system_info["disk_usage"].items():
                disk_info.append([
                    drive,
                    info.get("total_gb", "N/A"),
                    info.get("free_gb", "N/A"),
                    info.get("used_percent", "N/A")
                ])
        
        df_disk = pd.DataFrame(disk_info, columns=["Drive", "Total (GB)", "Livre (GB)", "Usado (%)"])
        
        # Escrever nas planilhas
        df_basic.to_excel(writer, sheet_name="Sistema", index=False, startrow=0)
        
        if not df_disk.empty:
            df_disk.to_excel(writer, sheet_name="Sistema", index=False, startrow=len(df_basic) + 3)
    
    def _create_programs_sheet(self, programs_data: Dict[str, Any], writer):
        """Cria aba de análise de programas"""
        
        # Programas por tamanho
        if "programs_by_size" in programs_data:
            programs_list = []
            for program in programs_data["programs_by_size"]:
                programs_list.append([
                    program.get("display_name", "N/A"),
                    program.get("size_mb", "N/A"),
                    program.get("install_date", "N/A"),
                    program.get("last_used", "N/A"),
                    program.get("version", "N/A")
                ])
            
            df_programs = pd.DataFrame(programs_list, columns=[
                "Programa", "Tamanho (MB)", "Data Instalação", "Último Uso", "Versão"
            ])
            df_programs.to_excel(writer, sheet_name="Programas", index=False)
        
        # Programas não utilizados
        if "unused_programs" in programs_data:
            unused_list = []
            for program in programs_data["unused_programs"]:
                unused_list.append([
                    program.get("display_name", "N/A"),
                    program.get("size_mb", "N/A"),
                    program.get("install_date", "N/A"),
                    program.get("version", "N/A")
                ])
            
            df_unused = pd.DataFrame(unused_list, columns=[
                "Programa", "Tamanho (MB)", "Data Instalação", "Versão"
            ])
            df_unused.to_excel(writer, sheet_name="Programas Não Utilizados", index=False)
    
    def _create_edge_sheet(self, edge_data: Dict[str, Any], writer):
        """Cria aba de análise do Edge"""
        
        # Top sites
        if "top_sites" in edge_data:
            sites_list = []
            for site in edge_data["top_sites"]:
                sites_list.append([
                    site.get("domain", "N/A"),
                    site.get("visit_count", "N/A"),
                    site.get("last_visit", "N/A")
                ])
            
            df_sites = pd.DataFrame(sites_list, columns=["Site", "Visitas", "Última Visita"])
            df_sites.to_excel(writer, sheet_name="Edge Analysis", index=False)
        
        # Estatísticas
        stats_data = []
        if "history_stats" in edge_data:
            stats = edge_data["history_stats"]
            stats_data.append(["Total de Entradas", stats.get("total_entries", "N/A")])
            stats_data.append(["Sites Únicos", stats.get("unique_sites", "N/A")])
            stats_data.append(["Período", stats.get("date_range", "N/A")])
        
        if "productivity_analysis" in edge_data:
            prod = edge_data["productivity_analysis"]
            stats_data.append(["Sites Desenvolvimento (%)", prod.get("dev_sites_percentage", "N/A")])
            stats_data.append(["Sites Sociais (%)", prod.get("social_sites_percentage", "N/A")])
        
        if stats_data:
            df_stats = pd.DataFrame(stats_data, columns=["Estatística", "Valor"])
            
            # Adicionar abaixo dos sites
            start_row = len(edge_data.get("top_sites", [])) + 3
            df_stats.to_excel(writer, sheet_name="Edge Analysis", index=False, startrow=start_row)
    
    def _create_bloatware_sheet(self, bloatware_list: List[Dict[str, Any]], writer):
        """Cria aba de bloatware"""
        
        bloatware_data = []
        for app in bloatware_list:
            bloatware_data.append([
                app.get("name", "N/A"),
                app.get("category", "N/A"),
                "Sim" if app.get("safe_to_remove", False) else "Verificar",
                app.get("description", "N/A")
            ])
        
        df_bloatware = pd.DataFrame(bloatware_data, columns=[
            "Aplicativo", "Categoria", "Seguro Remover", "Descrição"
        ])
        df_bloatware.to_excel(writer, sheet_name="Bloatware", index=False)
    
    def _apply_formatting(self, filepath: Path):
        """Aplica formatação ao arquivo Excel"""
        
        if not EXCEL_AVAILABLE:
            return
        
        try:
            # Carregar workbook
            wb = Workbook()
            wb = pd.ExcelFile(filepath).book if hasattr(pd.ExcelFile(filepath), 'book') else None
            
            if wb is None:
                return
            
            # Aplicar formatação básica
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'), 
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Formatação dos cabeçalhos
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                    cell.border = border
                
                # Auto-ajuste das colunas
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(filepath)
            
        except Exception as e:
            logger.warning(f"Erro ao aplicar formatação Excel: {e}")
    
    async def export_recommendations_report(self, recommendations: Dict[str, Any], filename: str = None) -> str:
        """Exporta relatório de recomendações em Excel"""
        
        if not EXCEL_AVAILABLE:
            raise ImportError("Bibliotecas Excel não disponíveis. Instale: pip install pandas openpyxl")
        
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"optimization_recommendations_{timestamp}.xlsx"
        
        filepath = self.output_dir / filename
        
        try:
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                
                # Criar abas para cada categoria
                categories = ["performance", "cleanup", "development", "security"]
                
                for category in categories:
                    if category in recommendations and len(recommendations[category]) > 0:
                        self._create_recommendations_sheet(
                            recommendations[category], 
                            category.title(), 
                            writer
                        )
            
            self._apply_formatting(filepath)
            
            logger.info(f"Relatório de recomendações Excel exportado para: {filepath}")
            return str(filepath)
            
        except Exception as e:
            logger.error(f"Erro ao exportar recomendações Excel: {e}")
            raise
    
    def _create_recommendations_sheet(self, recommendations: List[Dict], category: str, writer):
        """Cria aba de recomendações por categoria"""
        
        rec_data = []
        for rec in recommendations:
            rec_data.append([
                rec.get("priority", "N/A"),
                rec.get("category", "N/A"),
                rec.get("description", "N/A"),
                rec.get("action", "N/A")
            ])
        
        df_rec = pd.DataFrame(rec_data, columns=[
            "Prioridade", "Subcategoria", "Descrição", "Ação"
        ])
        df_rec.to_excel(writer, sheet_name=category, index=False)