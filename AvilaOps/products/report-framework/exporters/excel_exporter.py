"""
ÁVILA REPORT FRAMEWORK - EXPORTADOR EXCEL
==========================================
Gerador de relatórios em formato Excel com formatação avançada
"""

import os
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference, LineChart, PieChart
from config import get_timestamp, REPORT_TYPES
from logger import log_export_action

class ExcelExporter:
    """Exportador para formato Excel com formatação profissional"""

    def __init__(self):
        self.extension = ".xlsx"
        self.wb = None

    def export(self, data, report_type, filename=None):
        """Exportar dados para Excel"""
        try:
            log_export_action("excel", "arquivo", "started")

            if not filename:
                timestamp = get_timestamp("filename")
                filename = f"avila_report_{report_type}_{timestamp}{self.extension}"

            # Criar workbook
            self.wb = Workbook()
            self.wb.remove(self.wb.active)  # Remove worksheet padrão

            # Gerar planilhas baseado no tipo de relatório
            self._create_cover_sheet(data, report_type)
            self._create_summary_sheet(data, report_type)

            if report_type == "financial":
                self._create_financial_sheets(data)
            elif report_type == "projects":
                self._create_projects_sheets(data)
            elif report_type == "performance":
                self._create_performance_sheets(data)
            elif report_type == "governance":
                self._create_governance_sheets(data)
            else:
                self._create_generic_data_sheet(data)

            # Salvar arquivo
            filepath = os.path.join("exports", filename)
            os.makedirs("exports", exist_ok=True)
            self.wb.save(filepath)

            log_export_action("excel", filepath, "completed")
            return filepath

        except Exception as e:
            log_export_action("excel", "arquivo", "error")
            raise e

    def _create_cover_sheet(self, data, report_type):
        """Criar capa do relatório"""
        ws = self.wb.create_sheet("📋 Capa", 0)
        report_info = REPORT_TYPES.get(report_type, REPORT_TYPES["custom"])
        timestamp = get_timestamp("br")

        # Título principal
        ws['B2'] = f"🏛️ ÁVILA FRAMEWORK"
        ws['B2'].font = Font(size=24, bold=True, color="2F5597")

        ws['B4'] = f"{report_info['icon']} {report_info['name']}"
        ws['B4'].font = Font(size=18, bold=True)

        # Informações do relatório
        info_data = [
            ("📅 Data/Hora:", timestamp),
            ("📊 Tipo:", report_info['name']),
            ("🔄 Frequência:", report_info['frequency']),
            ("💻 Sistema:", "Ávila Report Framework v1.0"),
            ("👥 Responsável:", "AvilaOps Team")
        ]

        row = 7
        for label, value in info_data:
            ws[f'B{row}'] = label
            ws[f'C{row}'] = value
            ws[f'B{row}'].font = Font(bold=True)
            row += 1

        # Resumo executivo
        ws['B13'] = "📊 RESUMO EXECUTIVO"
        ws['B13'].font = Font(size=14, bold=True)
        ws['B15'] = data.get('summary', 'Resumo não disponível')

        # Formatação geral
        for col in ['B', 'C', 'D', 'E']:
            ws.column_dimensions[col].width = 20

    def _create_summary_sheet(self, data, report_type):
        """Criar planilha de resumo"""
        ws = self.wb.create_sheet("📈 Resumo")

        # Cabeçalho
        ws['A1'] = "📈 MÉTRICAS PRINCIPAIS"
        ws['A1'].font = Font(size=16, bold=True)

        # Métricas
        if 'metrics' in data:
            row = 3
            headers = ['Métrica', 'Valor', 'Status']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

            row += 1
            for metric_name, metric_value in data['metrics'].items():
                ws.cell(row=row, column=1, value=metric_name)
                ws.cell(row=row, column=2, value=metric_value)
                ws.cell(row=row, column=3, value="✅")
                row += 1

        # Auto-ajustar colunas
        for column in ws.columns:
            max_length = max(len(str(cell.value) or "") for cell in column)
            column_letter = column[0].column_letter
            ws.column_dimensions[column_letter].width = min(max_length + 2, 30)

    def _create_financial_sheets(self, data):
        """Criar planilhas específicas para relatório financeiro"""
        # Planilha de Receitas
        ws_receitas = self.wb.create_sheet("💰 Receitas")
        self._add_financial_data(ws_receitas, "Receitas", data.get('receitas_data', []))

        # Planilha de Despesas
        ws_despesas = self.wb.create_sheet("💸 Despesas")
        self._add_financial_data(ws_despesas, "Despesas", data.get('despesas_data', []))

        # Planilha de Análise
        ws_analise = self.wb.create_sheet("📊 Análise")
        self._add_financial_analysis(ws_analise, data)

    def _add_financial_data(self, ws, title, data):
        """Adicionar dados financeiros"""
        ws['A1'] = f"💰 {title.upper()}"
        ws['A1'].font = Font(size=14, bold=True)

        if not data:
            ws['A3'] = "Dados não disponíveis"
            return

        # Exemplo de estrutura de dados
        sample_data = [
            {"categoria": "Vendas", "valor": 50000, "mes": "Novembro"},
            {"categoria": "Serviços", "valor": 30000, "mes": "Novembro"},
            {"categoria": "Consultoria", "valor": 20000, "mes": "Novembro"},
        ]

        # Cabeçalhos
        headers = ['Categoria', 'Valor', 'Mês', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

        # Dados
        for row, item in enumerate(sample_data, 4):
            ws.cell(row=row, column=1, value=item['categoria'])
            ws.cell(row=row, column=2, value=item['valor'])
            ws.cell(row=row, column=3, value=item['mes'])
            ws.cell(row=row, column=4, value="✅")

    def _add_financial_analysis(self, ws, data):
        """Adicionar análise financeira"""
        ws['A1'] = "📊 ANÁLISE FINANCEIRA"
        ws['A1'].font = Font(size=14, bold=True)

        # Resumo financeiro
        financial_summary = [
            ("Total Receitas", data.get('receitas_total', 'N/A')),
            ("Total Despesas", data.get('despesas_total', 'N/A')),
            ("Resultado Líquido", data.get('resultado', 'N/A')),
            ("Margem de Lucro", data.get('margem', 'N/A')),
        ]

        row = 3
        for item, value in financial_summary:
            ws.cell(row=row, column=1, value=item).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1

    def _create_projects_sheets(self, data):
        """Criar planilhas para relatório de projetos"""
        ws = self.wb.create_sheet("🏗️ Projetos")

        # Dados de exemplo
        projects_data = [
            {"nome": "Projeto Alpha", "status": "Em Andamento", "progresso": "75%", "prazo": "15/12/2025"},
            {"nome": "Projeto Beta", "status": "Concluído", "progresso": "100%", "prazo": "01/11/2025"},
            {"nome": "Projeto Gamma", "status": "Atrasado", "progresso": "45%", "prazo": "30/10/2025"},
        ]

        # Cabeçalho
        ws['A1'] = "🏗️ PROJETOS"
        ws['A1'].font = Font(size=14, bold=True)

        headers = ['Projeto', 'Status', 'Progresso', 'Prazo']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

        # Dados
        for row, project in enumerate(projects_data, 4):
            ws.cell(row=row, column=1, value=project['nome'])
            ws.cell(row=row, column=2, value=project['status'])
            ws.cell(row=row, column=3, value=project['progresso'])
            ws.cell(row=row, column=4, value=project['prazo'])

    def _create_performance_sheets(self, data):
        """Criar planilhas para relatório de performance"""
        ws = self.wb.create_sheet("🚀 Performance")

        ws['A1'] = "🚀 INDICADORES DE PERFORMANCE"
        ws['A1'].font = Font(size=14, bold=True)

        performance_data = [
            ("Eficiência", data.get('score_eficiencia', '85%')),
            ("Qualidade", data.get('score_qualidade', '92%')),
            ("Produtividade", data.get('score_produtividade', '78%')),
            ("Satisfação Cliente", data.get('satisfacao_cliente', '89%')),
        ]

        row = 3
        for indicator, score in performance_data:
            ws.cell(row=row, column=1, value=indicator).font = Font(bold=True)
            ws.cell(row=row, column=2, value=score)
            row += 1

    def _create_governance_sheets(self, data):
        """Criar planilhas para relatório de governança"""
        ws = self.wb.create_sheet("🏛️ Governança")

        ws['A1'] = "🏛️ COMPLIANCE E GOVERNANÇA"
        ws['A1'].font = Font(size=14, bold=True)

        governance_data = [
            ("Score Compliance", data.get('score_compliance', '95%')),
            ("Não Conformidades", data.get('nao_conformidades', '2')),
            ("Auditorias Pendentes", data.get('auditorias_pendentes', '0')),
            ("Riscos Altos", data.get('riscos_alto', '1')),
        ]

        row = 3
        for item, value in governance_data:
            ws.cell(row=row, column=1, value=item).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1

    def _create_generic_data_sheet(self, data):
        """Criar planilha genérica para outros tipos de relatório"""
        ws = self.wb.create_sheet("📋 Dados")

        ws['A1'] = "📋 DADOS DO RELATÓRIO"
        ws['A1'].font = Font(size=14, bold=True)

        # Adicionar todos os dados disponíveis
        row = 3
        for key, value in data.items():
            ws.cell(row=row, column=1, value=str(key).title()).font = Font(bold=True)
            ws.cell(row=row, column=2, value=str(value))
            row += 1
